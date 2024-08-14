#!/usr/bin/python3

import sys
import openpyxl
from enum import Enum
from openpyxl.styles import Font, Border, Side, Alignment
# cli args library.
from sys import argv

# target INPUT_LOG_PATH/INPUT_LOG_FILENAME to analysis.
INPUT_FILE = argv[1]

# The output xlsx file:
OUTPUT_FILE = argv[2]

ParsePassType = [
    "Initialize",
    "Commit",
    "Evaluate",
    "Multi-opening Proof"
]

# PARSESTATE
class ParseStateType(Enum):
    STATE_DEFAULT       = 0
    STATE_INITIALIZE    = 1
    STATE_COMMIT        = 2
    STATE_EVALUATE      = 3
    STATE_MOP           = 4


creat_proof_state_dict = {
    "Initialize":                       ParseStateType.STATE_INITIALIZE,
    "Commit":                           ParseStateType.STATE_COMMIT,
    "Evaluate":                         ParseStateType.STATE_EVALUATE,
    "Multi-opening Proof":              ParseStateType.STATE_MOP,
}


class Trace:
    def __init__(self, name, level=1):
        self.name = name
        self.level = level

    def push_name(self, name):
        self.name.append(name)

    def pop_name(self):
        if len(self.name) > 0:
            return self.name.pop()
        else:
            return None
    def clear_name(self):
        self.name = []

    def add_level(self, level):
        self.level = level


class Node:
    def __init__(self, name, value=None, children=None, part=None):
        self.name = name
        # Time values
        self.value = [] if value is None else value
        # Mem values.
        self.mem_value = [] if value is None else value
        self.children = [] if children is None else children
        self.part = [] if part is None else part

    def add_value(self, value):
        self.value.append(value)

    def add_time_and_mem_value(self, time, mem_size):
        self.value.append(time)
        self.mem_value.append(mem_size)


    def add_part(self, module):
        self.part.append(module.queue[module.start:module.end])

    def print_structure(self, depth=0):
        indent = '  ' * depth
        print(f"{indent}- {self.name}: {self.value}")
        for child in self.children:
            child.print_structure(depth + 1)
        if self.part:
            for (i, segment) in enumerate(self.part):
                #print(obj)
                for obj in segment:
                    print(f" {indent}- [{i}] {obj.name}: {obj.value}")

    def find_nodes_with_children(self):
        nodes_with_children = []
        if self.children:
            nodes_with_children.append(self)
        for child in self.children:
            nodes_with_children.extend(child.find_nodes_with_children())
        return nodes_with_children

    def find_nodes_with_part(self):
        nodes_with_part = []
        if self.part:
            nodes_with_part.append(self)

        if self.children:
            for child in self.children:
                nodes_with_part.extend(child.find_nodes_with_part())

        return nodes_with_part


def find_node_by_name(node, target_name):
    if node.name == target_name:
        return node

    for child in node.children:
        found_node = find_node_by_name(child, target_name)
        if found_node is not None:
            return found_node

    return None


# self.name：
#   MSM CPU、MSM GPU、FFT CPU
# self.flag：   
#   FFT：
#   0 -> real fft
#   1 -> belong to ifft
class ModuleObject:
    def __init__(self, name, value, k, gpu, index, mem_bytes):
        self.name = name
        self.value = value
        self.k = k
        self.gpu = gpu
        self.index = index
        self.mem_size = mem_bytes            # fft/msm mem_size column
        self.flag = 0


class ExtraModule:
    def __init__(self, name, queue=None):
        self.name = name
        self.queue = [] if queue is None else queue
        self.start = 0
        self.end = 0
        self.msm_cpu_num = 0
        self.msm_gpu_num = 0
        self.msm_all_num = 0
        self.fft_cpu_num = 0
        self.fft_gpu_num = 0
        self.fft_num = 0
        self.ifft_cpu_num = 0
        self.ifft_gpu_num = 0
        self.ifft_num = 0
        self.fft_all_num = 0
        self.msm_cpu_time = 0.0
        self.msm_gpu_time = 0.0
        self.msm_all_time = 0.0
        self.fft_cpu_time = 0.0
        self.fft_gpu_time = 0.0
        self.fft_time = 0.0
        self.ifft_cpu_time = 0.0
        self.ifft_gpu_time = 0.0
        self.ifft_time = 0.0
        self.fft_all_time = 0.0


    def enqueue(self, obj):
        self.queue.append(obj)

    def query_last_gpu(self):
        if self.queue:
            length = len(self.queue)
            return self.queue[length-1].gpu
        else:
            return None

    def query_last_value(self):
        if self.queue:
            length = len(self.queue)
            return self.queue[length-1].value
        else:
            return None

    def query_last_obj(self):
        if self.queue:
            length = len(self.queue)
            return self.queue[length-1]
        else:
            return None

    # def remove_last_value():
    #     if self.value:
    #         length = len(value)
    #         self.value.pop(length-1)

    # def remove_last_k():
    #     if self.k:
    #         length = len(k)
    #         self.k.pop(length-1)

    # def remove_last_gpu():
    #     if self.gpu:
    #         length = len(gpu)
    #         self.gpu.pop(length-1)


# level >= 2
# depth >= 2
def parse_time_pass(root_node, line_split, current_trace_position, msm_module, fft_module):
    print("\nparse_time_pass: ", root_node.name)
    root_node.print_structure()
    print("     - current_trace_position name: ", current_trace_position.name)
    print("     - current_trace_position level: ", current_trace_position.level)
    # my_dict = {line_split[2]: line_split[3]}
    # initialize_list.append(my_dict)

    # Parse extra module
    # **|time_name|k|time|num|Bytes|
    # ----|Time|**|MSM GPU|5|0.008074417|1|
    if line_split[2] == "**":
        time_name = line_split[3]
        time_k =  line_split[4]
        time_value = line_split[5]
        time_num = line_split[6]
        mem_bytes = line_split[7]
        find_node = find_node_by_name(root_node, time_name)

        if time_name == 'MSM CPU':
            # name, value, k, gpu
            obj = ModuleObject(time_name, time_value, time_k, False, time_num, mem_bytes)
            msm_module.enqueue(obj)
            msm_module.msm_cpu_num += 1
            msm_module.msm_all_num += 1
            msm_module.msm_cpu_time += float(time_value)
            msm_module.msm_all_time += float(time_value)
            msm_module.end += 1
            assert msm_module.end == len(msm_module.queue), "Error: MSM CPU msm_module.end == len(msm_module.value) failed!"

        elif time_name == 'MSM GPU':
            obj = ModuleObject(time_name, time_value,time_k, True, time_num, mem_bytes)
            msm_module.enqueue(obj)
            msm_module.msm_gpu_num += 1
            msm_module.msm_all_num += 1
            msm_module.msm_gpu_time += float(time_value)
            msm_module.msm_all_time += float(time_value)
            msm_module.end += 1
            assert msm_module.end == len(msm_module.queue), "Error: MSM GPU msm_module.end == len(msm_module.value) failed!"

        elif time_name == 'FFT CPU':
            obj = ModuleObject(time_name, time_value,time_k, False, int(time_num) - fft_module.ifft_num, mem_bytes)
            fft_module.enqueue(obj)
            fft_module.fft_cpu_num += 1
            fft_module.fft_num += 1
            fft_module.fft_all_num += 1
            fft_module.fft_cpu_time += float(time_value)
            fft_module.fft_all_time += float(time_value)
            fft_module.end += 1
            assert fft_module.end ==len(fft_module.queue), "Error: FFT CPU fft_module.end == len(fft_module.value) failed!"

        elif time_name == 'FFT GPU':
            obj = ModuleObject(time_name, time_value,time_k, True, int(time_num) - fft_module.ifft_num, mem_bytes)
            fft_module.enqueue(obj)
            fft_module.fft_gpu_num += 1
            fft_module.fft_num += 1
            fft_module.fft_all_num += 1
            fft_module.fft_gpu_time += float(time_value)
            fft_module.fft_all_time += float(time_value)
            fft_module.end += 1
            assert fft_module.end == len(fft_module.queue), "Error: FFT CPU fft_module.end == len(fft_module.value) failed!"

        elif time_name == 'IFFT':

            fft_module.query_last_obj().name = fft_module.query_last_obj().name + " for IFFT"
            fft_module.query_last_obj().index = time_num

            # CPU
            if fft_module.query_last_gpu() == 0:
                value = float(fft_module.query_last_value())
                fft_module.fft_cpu_num -= 1
                fft_module.ifft_cpu_num += 1
                fft_module.fft_cpu_time -= value
                fft_module.ifft_cpu_time += float(time_value)
                obj = ModuleObject(time_name + " CPU", time_value,time_k, False, time_num, mem_bytes)
                fft_module.enqueue(obj)

            # GPU
            elif fft_module.query_last_gpu() == 1:
                value = float(fft_module.query_last_value())
                fft_module.fft_gpu_num -= 1
                fft_module.ifft_gpu_num += 1
                fft_module.fft_gpu_time -= value
                fft_module.ifft_gpu_time += float(time_value)
                obj = ModuleObject(time_name + " GPU", time_value,time_k, True, time_num, mem_bytes)
                fft_module.enqueue(obj)

            fft_module.fft_num -= 1
            fft_module.fft_all_num -= 1
            fft_module.ifft_num += 1
            fft_module.fft_all_num += 1

            fft_module.fft_all_time -= value
            fft_module.ifft_time += float(time_value)
            fft_module.fft_all_time += float(time_value)
            fft_module.end += 1
            assert fft_module.end == len(fft_module.queue), "Error: IFFT fft_module.end == len(fft_module.queue) failed!"

    # Parse module name part
    #     ----|Evaluate|Evaluate advice poly|----
    #     ----|Time|==|Evaluate advice poly|0.00318416|1Bytes|
    else:
        time_depth = line_split[2].count('=')
        time_name = line_split[3]
        print("     - time_depth: ", time_depth)
        print("     - time_name: ", time_name)

        if time_name == current_trace_position.name[current_trace_position.level - 1]:
            old_name = current_trace_position.name[current_trace_position.level - 1]
            current_trace_position.level = current_trace_position.level -  1
            current_trace_position.pop_name()
            print("     - new current_trace_position name: ", current_trace_position.name)
            print("     - new current_trace_position level: ", current_trace_position.level)

        assert time_depth == current_trace_position.level + 1, "depth should be equal to current_trace_position.level time_name: {}, \
                        time_depth: {}, current_trace_position.level: {}".format(time_name, time_depth, current_trace_position.level)

        for i, name in enumerate(current_trace_position.name):
            # 在循环体中执行操作
            print(i, name)
            find_node = find_node_by_name(root_node, name)
            if find_node == None:
                node = Node(name)
                parent_name = current_trace_position.name[i - 1]
                find_parent_node = find_node_by_name(root_node, parent_name)
                #create
                find_parent_node.children.append(node)

        find_node = find_node_by_name(root_node, time_name)
        print("     - find_node: ", find_node)

        if find_node == None:
            # create
            node = Node(time_name)
            # node.add_value(line_split[4])
            node.add_time_and_mem_value(line_split[4], line_split[5])
            #  todo add line_split[5]?
            parent_name = current_trace_position.name[time_depth - 2]
            print("     - parent_name: ", parent_name)
            find_parent_node = find_node_by_name(root_node, parent_name)
            print("     - find_parent_node: ", find_parent_node)
            assert find_parent_node != None, "Error: find_parent_node == None"
            find_parent_node.children.append(node)

            # add part
            if msm_module.start < msm_module.end:
                node.add_part(msm_module)
                msm_module.start = msm_module.end
            if fft_module.start < fft_module.end:
                node.add_part(fft_module)
                fft_module.start = fft_module.end

        else:
            # append
            # find_node.add_value(line_split[4])
            find_node.add_time_and_mem_value(line_split[4], line_split[5])
            # add part
            if msm_module.start < msm_module.end:
                find_node.add_part(msm_module)
                msm_module.start = msm_module.end
            if fft_module.start < fft_module.end:
                find_node.add_part(fft_module)
                fft_module.start = fft_module.end

    return root_node


def parse_state(line_split, current_trace_position):

    count = len(line_split)
    level = 0
    print("count: ", count)
    print("line_split: ", line_split)
    current_trace_position.clear_name()
    for i in range(2, count):
        if line_split[i] != "----":
            level = level + 1
            current_trace_position.push_name(line_split[i]);
            # name = line_split[i];
        else:
            break;

    current_trace_position.add_level(level)


    print("pasrs_state current_trace_position name: ", current_trace_position.name)
    print("pasrs_state current_trace_position level: ", current_trace_position.level)

    return current_trace_position


def parse_time_all_pass(create_proof_sum, line_split):

    create_proof_sum.append(line_split[2])
    create_proof_sum.append(line_split[3])
    print("parse_time_all_pass ", line_split[2], line_split[3])

def parse_input_file():

    print("\n\nparse_input_file")
    origin_file_lines = []
    input_file = INPUT_FILE

    # Read lines
    with open(input_file) as f:
        for line in f.readlines():
            origin_file_lines.append(line.strip('\n'))

    # print(origin_file_lines[0])
    # print(origin_file_lines[1])
    # print(origin_file_lines[2])
    # print(origin_file_lines[3])

    line_num = 0
    creat_proof_num = 0
    parse_state_type = ParseStateType.STATE_DEFAULT


    #init
    create_proof_list = []
    initialize_list = []
    commit_list = []
    evaluate_list = []
    mop_list = []
    current_state_position = "Default"
    current_trace_position = Trace("Default")

    node1 = Node(ParsePassType[0])
    node2 = Node(ParsePassType[1])
    node3 = Node(ParsePassType[2])
    node4 = Node(ParsePassType[3])

    msm_module = ExtraModule("MSM")
    fft_module = ExtraModule("FFT")

    create_proof_sum = []
    all_create_proof_list = []


    # Parse lines
    for line in origin_file_lines:
        line_num += 1
        if line == "":
            continue
        else:
            index = line.find("----|")

            if (index != -1):
                new_line = line[index:]

                line_split = new_line.strip().split("|")
                # Found creat proof start
                if len(line_split) == 1:
                    continue
                elif line_split[0] == "----":
                    # title
                    if line_split[1] == "Profiling":

                        if line_split[2] == "Creat proof start":
                            creat_proof_num += 1
                            print("creat proof start, count: ", creat_proof_num)

                        elif line_split[2] == "Creat proof end":
                            print("creat proof end, count: ", creat_proof_num)
                            break

                        else:
                            current_state_position = line_split[2]
                            parse_state_type = creat_proof_state_dict[current_state_position]
                            current_trace_position = parse_state(line_split, current_trace_position)

                    # time
                    elif line_split[1] == "Time":
                        if  parse_state_type == ParseStateType.STATE_INITIALIZE:
                            parse_time_pass(node1, line_split, current_trace_position, msm_module, fft_module)
                        elif  parse_state_type == ParseStateType.STATE_COMMIT:
                            parse_time_pass(node2, line_split, current_trace_position, msm_module, fft_module)
                        elif  parse_state_type == ParseStateType.STATE_EVALUATE:
                            parse_time_pass(node3, line_split, current_trace_position, msm_module, fft_module)
                        elif  parse_state_type == ParseStateType.STATE_MOP:
                            parse_time_pass(node4, line_split, current_trace_position, msm_module, fft_module)
                        elif  parse_state_type == ParseStateType.STATE_DEFAULT:
                            print("Warning: parse time in ParseStateType.STATE_DEFAULT")

                    elif line_split[1] == "Time All":
                        parse_time_all_pass(create_proof_sum, line_split)


    create_proof_list = [node1, node2, node3, node4]
    all_create_proof_list = [create_proof_list, msm_module, fft_module, create_proof_sum]
    print("\n\nparse_input_file draw")
    node1.print_structure()
    node2.print_structure()
    node3.print_structure()
    node4.print_structure()

    return all_create_proof_list




from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
# columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
# TITLE_NAME1 = ['Module name', 'Time(s)']
# TITLE_NAME2 = ['Module name', 'index', 'K','Time(s)']
# TITLE_NAME3 = ['Index', 'Module name', 'Total number', 'Total time(us)']
class LayoutConfig:
    def __init__(self, title_font=None, context_font=None, alignment=None, border=None):
        self.title_font =  Font(name='Arial', size=11, bold=True, italic=True)
        self.context_font = Font(name='Arial', size=11, bold=True, italic=True)
        self.alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        self.first_column_width = 60
        self.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        # for summary sheet
        self.title_name1= ['Index', 'Module name', 'Total time(s)']
        self.title_name2 = ['Index', 'Module name', 'Total number', 'Total time(s)']
        # for detail sheet, including time/mem
        self.title_name3 = ['Module name', 'Time(s)', 'Mem(Byte)']
        self.title_name4 = ['Module name', 'index', 'K','Time(s)', 'Mem(Byte)']


def config_output_file():

    workbook = openpyxl.Workbook()

    output_file_name = OUTPUT_FILE

    # 创建字体对象
    title_font = Font(name='Arial', size=11, bold=True, italic=False)
    context_font = Font(name='Times New Roman', size=10, bold=False, italic=False)
    alignment = Alignment(horizontal='center', vertical='center')
    # 创建边框对象
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    layout_config = LayoutConfig()
    layout_config.title_font = title_font
    layout_config.context_font = context_font
    layout_config.alignment = alignment
    layout_config.border = border

    return workbook, output_file_name, layout_config


# format
def gen_title_cell_format(sheet, layout_config, column, row, format_type):

    cell = sheet[column + str(row)]
    cell.font = layout_config.title_font
    cell.border = layout_config.border
    cell.alignment = layout_config.alignment
    column_width = len(str(cell.value))
    adjusted_width = (column_width + 2) * 1.6
    sheet.column_dimensions[column].width = adjusted_width
    # sheet.column_dimensions[column_letter].auto_size  = True          #max(sheet1.column_dimensions[column_letter].width, len(str(cell.value)))

    if format_type == 0:
        pass
    elif format_type == 1:
        # yellow
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')
    elif format_type == 2:
        # red
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')


    return


def gen_context_cell_format(sheet, layout_config, column, current_row_position, format_type):

    cell = sheet[column + str(current_row_position)]
    cell.font = layout_config.context_font
    cell.border = layout_config.border
    cell.alignment = layout_config.alignment

    return


def gen_context_area_format(sheet, layout_config, columns, current_row_position, row_length, format_type):

    for i in range(current_row_position, (current_row_position + row_length)):
        for column in columns:
            gen_context_cell_format(sheet, layout_config, column, i, format_type)

    return sheet, current_row_position


# title
def gen_table_title (   sheet,
                        layout_config,
                        columns,
                        current_row_position,
                        title_names,
                        title_type,
                        format_type,
                        node):

    # For summary sheet
    if title_type == 1:
        for (i, column) in enumerate(columns):
            sheet[column + str(current_row_position)] = title_names[0][i]
            gen_title_cell_format(sheet, layout_config, column, current_row_position, format_type)

        current_row_position += 1

    # For sub sheet
    elif title_type == 2:
        # table name
        sheet[columns[0] + str(current_row_position)] = node.name
        gen_title_cell_format(sheet, layout_config, layout_config.columns[0], current_row_position, format_type)        # table title

        current_row_position = current_row_position + 1

        start = 0
        end = 2 # add column for non merged one.

        if node.part:
            # module
            # Module name	Time(s)	Module name	K Time(s)
            for j in range(0,2):
                # start = j * 2
                # end = (j+1) * 2
                if j == 1:
                    start = 2
                    end = 6
                for (i, column) in enumerate(columns[start:end+1]):
                    sheet[column + str(current_row_position)] = title_names[0][i] if j == 0 else title_names[1][i]
                    cell = sheet[column + str(current_row_position)]
                    gen_title_cell_format(sheet, layout_config, column, current_row_position, format_type)
        else:
            for (i, column) in enumerate(columns[start:end+1]):
                sheet[column + str(current_row_position)] = title_names[0][i]
                cell = sheet[column + str(current_row_position)]
                gen_title_cell_format(sheet, layout_config, column, current_row_position, format_type)

        current_row_position = current_row_position + 1

    return sheet, current_row_position


def gen_layout_summary_sheet(workbook, layout_config, all_create_proof_list):

    create_proof_list = all_create_proof_list[0]
    msm_module = all_create_proof_list[1]
    fft_module = all_create_proof_list[2]
    create_proof_sum = all_create_proof_list[3]

    sheet = workbook.active
    sheet.title = "Summary"

    current_row_position = 1
    between_segment_row_num = 3

    ########## segment 1
    length = len(create_proof_list)
    # gen title
    sheet, current_row_position = gen_table_title(sheet, layout_config, layout_config.columns[0:3], current_row_position, [layout_config.title_name1], 1, 1, None)

    # gen context
    for i in range(0, length):
        root_node = create_proof_list[i]
        name = ParsePassType[i]
        value = root_node.value

        sheet[layout_config.columns[0] + str(current_row_position + i)] = i+1
        sheet[layout_config.columns[1] + str(current_row_position + i)] = name
        if len(value) > 0:
            sheet[layout_config.columns[2] + str(current_row_position + i)] = float(value[0])
        else:
            sheet[layout_config.columns[2] + str(current_row_position + i)] = ""

    if create_proof_sum:
        name = create_proof_sum[0]
        value = create_proof_sum[1]
        sheet[layout_config.columns[0] + str(current_row_position + length)] = "total"
        sheet[layout_config.columns[1] + str(current_row_position + length)] = name

        if value:
            sheet[layout_config.columns[2] + str(current_row_position + length)] = float(value)
        else:
            sheet[layout_config.columns[2] + str(current_row_position + length)] = ""

        length += 1

    # gen format
    sheet, current_row_position = gen_context_area_format(sheet, layout_config, layout_config.columns[0:3], current_row_position, length, 1)
    current_row_position += length

    current_row_position += between_segment_row_num

    ########## segment 2    msm
    # gen title
    sheet, current_row_position = gen_table_title(sheet, layout_config, layout_config.columns[0:4], current_row_position, [layout_config.title_name2], 1, 1, None)

    data = [
        ("MSM CPU", msm_module.msm_cpu_num, msm_module.msm_cpu_time),
        ("MSM GPU", msm_module.msm_gpu_num, msm_module.msm_gpu_time),
        ("MSM", msm_module.msm_all_num, msm_module.msm_all_time)
    ]

    for i, row_data in enumerate(data):
        row_index = current_row_position + i
        sheet[layout_config.columns[0] + str(row_index)] = i + 1
        sheet[layout_config.columns[1] + str(row_index)] = row_data[0]
        sheet[layout_config.columns[2] + str(row_index)] = row_data[1]
        sheet[layout_config.columns[3] + str(row_index)] = row_data[2]

    length = 3
    # gen context

    # format
    sheet, current_row_position = gen_context_area_format(sheet, layout_config, layout_config.columns[0:4], current_row_position, length, 0)
    current_row_position += length
    current_row_position += between_segment_row_num

    ########## segment 3    fft
    # gen context - fft
    sheet, current_row_position = gen_table_title(sheet, layout_config, layout_config.columns[0:4], current_row_position, [layout_config.title_name2], 1, 1, None)

    data = [
        ("FFT CPU", fft_module.fft_cpu_num, fft_module.fft_cpu_time),
        ("FFT GPU", fft_module.fft_gpu_num, fft_module.fft_gpu_time),
        ("FFT", fft_module.fft_gpu_num, fft_module.fft_gpu_time),
        ("IFFT CPU", fft_module.ifft_cpu_num, fft_module.ifft_cpu_time),
        ("IFFT GPU", fft_module.ifft_gpu_num, fft_module.ifft_gpu_time),
        ("IFFT", fft_module.ifft_num, fft_module.ifft_time),
        ("FFT + IFFT ALL", fft_module.fft_all_num, fft_module.fft_all_time)
    ]

    for i, row_data in enumerate(data):
        row_index = current_row_position + i
        sheet[layout_config.columns[0] + str(row_index)] = i + 1
        sheet[layout_config.columns[1] + str(row_index)] = row_data[0]
        sheet[layout_config.columns[2] + str(row_index)] = row_data[1]
        sheet[layout_config.columns[3] + str(row_index)] = row_data[2]

    length = 7

    # format
    sheet, current_row_position = gen_context_area_format(sheet, layout_config, layout_config.columns[0:4], current_row_position, length, 0)
    current_row_position += length

    current_row_position += between_segment_row_num

    return workbook


def gen_table_context(sheet, layout_config, columns, current_row_position, format_type, node):


    length = len(node.children[0].value)

    for child_node in node.children:
        name = child_node.name
        value = child_node.value
        assert length == len(value), "children node value length not the same"

    for j in range(length):
        for child_node in node.children:
            name = child_node.name
            value = child_node.value
            mem_val = child_node.mem_value
            context = [name, float(value[j]), mem_val[j]] # assign value: [name,time,mem_size]
            for (i, column) in enumerate(columns):           #enumerate(columns[:2]):
                sheet[column + str(current_row_position)] = context[i]
                gen_context_cell_format(sheet, layout_config, column, current_row_position, format_type)
            current_row_position = current_row_position + 1
        if j < length -1:
            current_row_position = current_row_position + 1

    return sheet, current_row_position


def gen_table_context_with_module(sheet, layout_config, columns, current_row_position, format_type, node):

    loop_count = len(node.value)
    print(loop_count, node.name)
    #assert loop_count == len(node.value), "Error: [gen_table_context_with_module] loop_count == len(node.value) is failed!"

    part_count = len(node.part)
    assert loop_count != 0, "Error: [gen_table_context_with_module] part_count != 0"
    print(part_count, node.name)

    each_loop_part_count = int(part_count / loop_count)

    each_loop_len = 0
    for i in range(0, each_loop_part_count):
        each_loop_len = each_loop_len + len(node.part[i])

    assert each_loop_len != 0, "Error: [gen_table_context_with_module] each_loop_len != 0"
    total_count = each_loop_len * loop_count
    tatal_row_num = total_count

    colunm_start = 0
    colunm_end = 0
    old_current_row_position = current_row_position

    # gen merge pass
    for i in range(0, loop_count):
        # 1. module name.
        merge_areaA1 = columns[0]+ str(current_row_position)
        merge_areaA2 = columns[0]+ str(current_row_position + each_loop_len - 1)

        merge_areaA = merge_areaA1 + ":" + merge_areaA2

        sheet.merge_cells(merge_areaA)
        sheet[merge_areaA1] = node.name
        gen_context_cell_format(sheet, layout_config, columns[0], current_row_position, format_type)

        # 2. Time
        merge_areaB1 = columns[1]+ str(current_row_position)
        merge_areaB2 = columns[1]+ str(current_row_position + each_loop_len - 1)
        merge_areaB = merge_areaB1 + ":" + merge_areaB2

        sheet.merge_cells(merge_areaB)
        sheet[merge_areaB1] = float(node.value[i])

        gen_context_cell_format(sheet, layout_config, columns[1], current_row_position, format_type)

        # TODO: how to assign value for the mem_size?~
        # # 3. Mem size for the merge cell
        # merge_areaC1 = columns[2]+ str(current_row_position)
        # merge_areaC2 = columns[2]+ str(current_row_position + each_loop_len - 1)
        # merge_areaC = merge_areaC1 + ":" + merge_areaC2
        # sheet.merge_cells(merge_areaC)
        # # sheet[merge_areaC1] = node.mem_value[i]
        # sheet[merge_areaC1] = "111"
        #
        # gen_context_cell_format(sheet, layout_config, columns[2], current_row_position, format_type)

        current_row_position = current_row_position + each_loop_len

    current_row_position = old_current_row_position


    # gen module pass
    for (j, part) in enumerate(node.part):
        for (i, obj) in enumerate(part):
            sheet[columns[2]+ str(current_row_position)] = obj.name
            gen_context_cell_format(sheet, layout_config, columns[2], current_row_position, format_type)

            sheet[columns[3]+ str(current_row_position)] = int(obj.index)
            gen_context_cell_format(sheet, layout_config, columns[3], current_row_position, format_type)

            sheet[columns[4]+ str(current_row_position)] = int(obj.k)
            gen_context_cell_format(sheet, layout_config, columns[4], current_row_position, format_type)

            sheet[columns[5]+ str(current_row_position)] = float(obj.value)
            gen_context_cell_format(sheet, layout_config, columns[5], current_row_position, format_type)

            # sheet[columns[6]+ str(current_row_position)] = int(obj.mem_size)
            # fft/msm mem_size column
            sheet[columns[6]+ str(current_row_position)] = obj.mem_size
            gen_context_cell_format(sheet, layout_config, columns[6], current_row_position, format_type)

            current_row_position = current_row_position + 1

    return sheet, current_row_position


def gen_layout_sub_sheet(workbook, layout_config, root_node):

    print("\n\n==> gen_layout_sub_sheet")

    sheet_title = root_node.name
    sheet = workbook.create_sheet(sheet_title)

    current_row_position = 1

    # calculate table number
    nodes_with_children = root_node.find_nodes_with_children()

    # gen layout
    for (i, node) in enumerate(nodes_with_children):
        sheet, current_row_position = gen_table_title(sheet, layout_config, layout_config.columns, current_row_position, [layout_config.title_name3, layout_config.title_name4], 2, 1, node)  #gen_table_title(sheet, layout_config, current_row_position, node)
        sheet, current_row_position = gen_table_context(sheet, layout_config, layout_config.columns[:3], current_row_position, 0, node)
        sheet.column_dimensions[layout_config.columns[0]].width = layout_config.first_column_width
        current_row_position = current_row_position + 3


    # calculate module table number
    nodes_with_part = root_node.find_nodes_with_part()

    #print("nodes_with_part: ", nodes_with_part)
    for (i, node) in enumerate(nodes_with_part):
        sheet, current_row_position = gen_table_title(sheet, layout_config, layout_config.columns, current_row_position, [layout_config.title_name3, layout_config.title_name4], 2, 1, node) #gen_table_title(sheet, layout_config, current_row_position, node)
        sheet, current_row_position = gen_table_context_with_module(sheet, layout_config, layout_config.columns, current_row_position, 0, node)
        sheet.column_dimensions[layout_config.columns[0]].width = layout_config.first_column_width
        current_row_position = current_row_position + 3

    return workbook


def gen_layout(workbook, layout_config, all_create_proof_list):


    workbook = gen_layout_summary_sheet(workbook, layout_config, all_create_proof_list);
    create_proof_list = all_create_proof_list[0]
    for (i, root_node) in enumerate(create_proof_list):
        workbook = gen_layout_sub_sheet(workbook, layout_config, root_node);

    return workbook


def generate_output_file(all_create_proof_list):

    print("generate_file")

    workbook, output_file_name, layout_config = config_output_file()

    gen_layout(workbook, layout_config, all_create_proof_list)

    workbook.save(output_file_name)

    return


def main():

    all_create_proof_list = parse_input_file()

    if all_create_proof_list:
        generate_output_file(all_create_proof_list)

    return


# debug command
#  ./halo2_proofs/tool/profiling_analysis.py ./logs/test_1 ./logs/test_1.xlsx >> logs/debug_python
if __name__== "__main__":
    main()
